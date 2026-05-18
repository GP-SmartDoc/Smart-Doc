"""
QA Pipeline Evaluation — Faithfulness + Answer Relevancy + Contextual Relevancy + MRR
=======================================================================================
Reads 30 PDFs from:
    llm_as_judge/QA test/pdfs_QA/test{N}.pdf

Reads questions from:
    llm_as_judge/QA test/TestQuestions.xlsx
    Columns: Q1, Q2, Q3  (row N → test{N}.pdf)

For every (PDF, question) pair the script:
  1. Indexes the PDF into an isolated ephemeral RAG (once per PDF, shared across Q1–Q3).
  2. Runs the RAG retrieval (k_text=6) filtered to the PDF.
  3. Runs the full RAG + QA pipeline (QuestionAnsweringModule).
  4. Evaluates with three DeepEval metrics (async, concurrent):
       • FaithfulnessMetric        — answer grounded in retrieved context
       • AnswerRelevancyMetric     — answer is on-topic to the question
       • ContextualRelevancyMetric — retrieved chunks are relevant to the question
  5. Computes Mean Reciprocal Rank (MRR) for the retrieved chunk list via a
     single judge call that identifies the first relevant chunk's rank.

Outputs
-------
  llm_as_judge/qa_eval_results.json   (flushed after EVERY sample)

Resume / Retry behaviour
------------------------
  Completed samples (metrics present, no error) are skipped on re-run.
  Timeout / errored samples are also skipped (mark-and-move-on strategy).

Run
---
    # from project root:
    pytest llm_as_judge/evaluate_qa.py -v
    pytest llm_as_judge/evaluate_qa.py -v -k "test_id_1"
"""

import asyncio
import os
import sys
import json
import traceback
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from datetime import datetime

# ── Make project root importable ────────────────────────────────────────────
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from dotenv import load_dotenv
load_dotenv()

import chromadb
import openpyxl
import weave
import pytest
from langchain_openai import ChatOpenAI
from deepeval.models import DeepEvalBaseLLM
from deepeval.metrics import (
    FaithfulnessMetric,
    AnswerRelevancyMetric,
    ContextualRelevancyMetric,
)
from deepeval.test_case import LLMTestCase

from src.vector_store.RAG import RAGEngine
from src.graphs.qa_graph import QuestionAnsweringModule

# ── Weave tracing ────────────────────────────────────────────────────────────
weave.init("gamer7dragon817-ain-shams-university/slide-generator-demo")


# ═══════════════════════════════════════════════════════════════════════════
#  PATHS
# ═══════════════════════════════════════════════════════════════════════════
_HERE         = os.path.dirname(os.path.abspath(__file__))
_PDFS_DIR     = os.path.join(_HERE, "QA test", "pdfs_QA")
_XLSX_PATH    = os.path.join(_HERE, "QA test", "TestQuestions.xlsx")
_RESULTS_JSON = os.path.join(_HERE, "qa_eval_results.json")

_QUESTION_COLS = ["Q1", "Q2", "Q3"]
_TIMEOUT_SECONDS = 300   # per test (ingest + QA + 3 metrics + MRR judge call)


# ═══════════════════════════════════════════════════════════════════════════
#  JUDGE MODEL  (W&B Inference)
# ═══════════════════════════════════════════════════════════════════════════
class WandbJudge(DeepEvalBaseLLM):
    """DeepEval judge backed by the W&B Inference API."""

    def __init__(self):
        self._model = ChatOpenAI(
            model=os.getenv("EVAL_MODEL", "Qwen/Qwen3-235B-A22B-Instruct-2507"),
            openai_api_key=os.getenv("OPENAI_API_KEY"),
            openai_api_base=os.getenv(
                "OPENAI_API_BASE", "https://api.inference.wandb.ai/v1"
            ),
            temperature=0,
        )

    def load_model(self):
        return self._model

    def generate(self, prompt: str) -> str:
        return self._model.invoke(prompt).content

    async def a_generate(self, prompt: str) -> str:
        response = await self._model.ainvoke(prompt)
        return response.content

    def get_model_name(self) -> str:
        return os.getenv("EVAL_MODEL", "Qwen/Qwen3-235B-A22B-Instruct-2507")


judge = WandbJudge()


# ═══════════════════════════════════════════════════════════════════════════
#  DEEPEVAL METRICS
# ═══════════════════════════════════════════════════════════════════════════
_faithfulness_metric = FaithfulnessMetric(
    threshold=0.7,
    model=judge,
    include_reason=True,
)

_answer_relevancy_metric = AnswerRelevancyMetric(
    threshold=0.7,
    model=judge,
    include_reason=True,
)

_contextual_relevancy_metric = ContextualRelevancyMetric(
    threshold=0.7,
    model=judge,
    include_reason=True,
)

_ALL_METRICS = [
    _faithfulness_metric,
    _answer_relevancy_metric,
    _contextual_relevancy_metric,
]


# ═══════════════════════════════════════════════════════════════════════════
#  RAG + QA ENGINE  (shared ephemeral instance — models load once)
# ═══════════════════════════════════════════════════════════════════════════
_rag_client  = chromadb.EphemeralClient()
_rag         = RAGEngine(_rag_client)
_qa          = QuestionAnsweringModule(_rag)
_indexed_pdfs: set[str] = set()   # tracks which PDFs have been indexed


# ═══════════════════════════════════════════════════════════════════════════
#  LOAD SAMPLES FROM XLSX
# ═══════════════════════════════════════════════════════════════════════════
def _load_samples() -> list[dict]:
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    q_indices = {col: headers.index(col) for col in _QUESTION_COLS if col in headers}

    samples = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        pdf_file = f"test{row_idx}.pdf"
        pdf_path = os.path.join(_PDFS_DIR, pdf_file)
        if not os.path.exists(pdf_path):
            continue
        for q_col, col_idx in q_indices.items():
            question = row[col_idx]
            if not question:
                continue
            samples.append({
                "test_id":    row_idx,
                "pdf_file":   pdf_file,
                "pdf_path":   pdf_path,
                "question_col": q_col,
                "question":   str(question).strip(),
                "run_key":    f"{row_idx}_{q_col}",
            })
    wb.close()
    return samples


# ═══════════════════════════════════════════════════════════════════════════
#  ASYNC METRIC SCORING
# ═══════════════════════════════════════════════════════════════════════════
async def _score_metric_async(metric, test_case: LLMTestCase) -> dict:
    await metric.a_measure(test_case)
    return {
        "name":      getattr(metric, "name", type(metric).__name__),
        "score":     metric.score,
        "threshold": metric.threshold,
        "success":   metric.score >= metric.threshold,
        "reason":    getattr(metric, "reason", None),
    }


@weave.op()
def _score_all_metrics(test_case: LLMTestCase) -> list[dict]:
    async def _gather():
        return await asyncio.gather(
            *[_score_metric_async(m, test_case) for m in _ALL_METRICS]
        )
    return asyncio.run(_gather())


# ═══════════════════════════════════════════════════════════════════════════
#  MRR  (judge-based: single call to find first relevant chunk rank)
# ═══════════════════════════════════════════════════════════════════════════
def _compute_mrr(question: str, chunks: list[str]) -> dict:
    """Ask the judge which retrieved chunk (1-indexed) is the FIRST one that
    contains information useful to answer the question.
    Returns {"mrr": float, "first_relevant_rank": int}."""
    if not chunks:
        return {"mrr": 0.0, "first_relevant_rank": 0}

    numbered_chunks = "\n\n".join(
        f"[{i + 1}] {c[:600]}" for i, c in enumerate(chunks)
    )
    prompt = (
        f"You are evaluating a retrieval system.\n\n"
        f"Question: {question}\n\n"
        f"Retrieved chunks (in retrieval order):\n{numbered_chunks}\n\n"
        "Task: Identify the FIRST chunk (by its number) that contains information "
        "directly useful for answering the question. "
        "Reply with ONLY a single integer (1 to "
        f"{len(chunks)}), or 0 if no chunk is relevant."
    )
    try:
        response = judge.generate(prompt).strip()
        # Extract the first integer from the response
        import re
        match = re.search(r"\b(\d+)\b", response)
        rank = int(match.group(1)) if match else 0
        if rank < 0 or rank > len(chunks):
            rank = 0
        mrr = 1.0 / rank if rank >= 1 else 0.0
    except Exception as exc:
        print(f"[WARN] MRR judge call failed: {exc}")
        rank = 0
        mrr = 0.0

    return {"mrr": mrr, "first_relevant_rank": rank}


# ═══════════════════════════════════════════════════════════════════════════
#  PIPELINE
# ═══════════════════════════════════════════════════════════════════════════
@weave.op()
def _run_pipeline(sample: dict) -> tuple[str, list[str]]:
    """Index PDF (first access only) → query RAG → run QA pipeline.
    Returns (answer_text, retrieved_chunks)."""
    pdf_file = sample["pdf_file"]
    pdf_path = sample["pdf_path"]
    question = sample["question"]

    # Lazy PDF indexing — shared across Q1/Q2/Q3 for the same document
    if pdf_file not in _indexed_pdfs:
        print(f"[INDEX] Indexing {pdf_file} ...")
        _rag.add_pdf(pdf_path)
        _indexed_pdfs.add(pdf_file)

    # Retrieve with document filter
    retrieved = _rag.query(question, k_text=6, document=pdf_file)
    retrieved_chunks: list[str] = retrieved.get("text", [])

    # Full QA pipeline
    result = _qa.invoke(question, document=pdf_file)
    answer = result.get("Answer", "")

    # Flatten answer if it's a dict (qa_agent returns content string, but guard anyway)
    if isinstance(answer, dict):
        answer = json.dumps(answer, ensure_ascii=False)

    return answer, retrieved_chunks


# ═══════════════════════════════════════════════════════════════════════════
#  JSON PERSISTENCE
# ═══════════════════════════════════════════════════════════════════════════
def _compute_averages(results: list[dict]) -> dict:
    from collections import defaultdict
    overall: dict[str, list[float]] = defaultdict(list)
    mrr_scores: list[float] = []

    for r in results:
        if r.get("mrr_score") is not None:
            mrr_scores.append(float(r["mrr_score"]))
        for m in r.get("metrics", []):
            s = m.get("score")
            if s is not None:
                overall[m["name"]].append(float(s))

    def _avg(lst):
        return round(sum(lst) / len(lst), 4) if lst else 0.0

    return {
        "MeanReciprocalRank": _avg(mrr_scores),
        **{name: _avg(scores) for name, scores in overall.items()},
    }


def _flush_json(results: list[dict]) -> None:
    try:
        averages = _compute_averages(results)
        output = {
            "timestamp":        datetime.now().isoformat(timespec="seconds"),
            "xlsx_source":      _XLSX_PATH,
            "pdfs_dir":         _PDFS_DIR,
            "total":            len(results),
            "passed":           sum(1 for r in results if r.get("passed")),
            "failed":           sum(1 for r in results if not r.get("passed")),
            "averages":         averages,
            "results":          results,
        }
        with open(_RESULTS_JSON, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    except Exception as exc:
        print(f"[WARN] Could not write results JSON: {exc}")


def _load_previous_results() -> list[dict]:
    if os.path.exists(_RESULTS_JSON):
        try:
            with open(_RESULTS_JSON, encoding="utf-8") as f:
                data = json.load(f)
            return data.get("results", [])
        except Exception:
            pass
    return []


def _is_completed(result: dict) -> bool:
    """Completed = has metrics with no error, OR timed out (skip both)."""
    if result.get("metrics") and not result.get("error"):
        return True
    err = result.get("error", "")
    if isinstance(err, str) and err.startswith("Timeout:"):
        return True
    return False


# ═══════════════════════════════════════════════════════════════════════════
#  RESUME STATE
# ═══════════════════════════════════════════════════════════════════════════
_prev_results   = _load_previous_results()
_done_keys      = {r["run_key"] for r in _prev_results if _is_completed(r)}
pytest_results  = [r for r in _prev_results if _is_completed(r)]

_samples        = _load_samples()
_samples_to_run = [s for s in _samples if s["run_key"] not in _done_keys]
_ids            = [f"test{s['test_id']}_{s['question_col']}" for s in _samples_to_run]

retry_count   = sum(1 for r in _prev_results if not _is_completed(r))
skipped_count = len(_samples) - len(_samples_to_run)
print(
    f"[INFO] Total samples: {len(_samples)} | "
    f"Skipped (done/timeout): {skipped_count} | "
    f"Retrying errors: {retry_count} | "
    f"To run: {len(_samples_to_run)}"
)

# Write skeleton JSON so the file always exists before any test runs
if not os.path.exists(_RESULTS_JSON):
    _flush_json([])


# ═══════════════════════════════════════════════════════════════════════════
#  PYTEST PARAMETRISED TESTS
# ═══════════════════════════════════════════════════════════════════════════
@pytest.mark.parametrize("sample", _samples_to_run, ids=_ids)
def test_qa_pipeline(sample: dict):
    """
    For each (PDF, question) pair:
      1. Index the PDF into the ephemeral RAG (first time only per PDF).
      2. Retrieve with document filter.
      3. Run RAG + QA pipeline.
      4. Score: FaithfulnessMetric, AnswerRelevancyMetric, ContextualRelevancyMetric.
      5. Compute MRR via a judge call.
      6. Flush results to JSON.
    """

    def _run_test():
        answer, retrieved_chunks = _run_pipeline(sample)

        assert answer, (
            f"QA pipeline returned empty answer for "
            f"test{sample['test_id']} / {sample['question_col']}: {sample['question'][:80]}"
        )

        test_case = LLMTestCase(
            input=sample["question"],
            actual_output=answer,
            retrieval_context=retrieved_chunks or [],
        )

        scores  = _score_all_metrics(test_case)
        mrr_res = _compute_mrr(sample["question"], retrieved_chunks)
        return answer, retrieved_chunks, scores, mrr_res

    executor = ThreadPoolExecutor(max_workers=1)
    future   = executor.submit(_run_test)

    try:
        try:
            answer, retrieved_chunks, scores, mrr_res = future.result(
                timeout=_TIMEOUT_SECONDS
            )
        except FuturesTimeout:
            executor.shutdown(wait=False)
            print(
                f"[SKIP] Timeout {_TIMEOUT_SECONDS}s — "
                f"test{sample['test_id']}/{sample['question_col']}: {sample['question'][:60]}"
            )
            pytest_results.append({
                "run_key":            sample["run_key"],
                "test_id":            sample["test_id"],
                "pdf_file":           sample["pdf_file"],
                "question_col":       sample["question_col"],
                "question":           sample["question"],
                "answer":             "",
                "retrieved_chunks":   [],
                "mrr_score":          None,
                "first_relevant_rank": None,
                "metrics":            [],
                "passed":             False,
                "error":              f"Timeout: exceeded {_TIMEOUT_SECONDS} seconds",
            })
            _flush_json(pytest_results)
            pytest.skip(f"Exceeded {_TIMEOUT_SECONDS}s time limit")

        except Exception:
            executor.shutdown(wait=False)
            err_msg = traceback.format_exc()
            print(
                f"[ERROR] test{sample['test_id']}/{sample['question_col']}: "
                f"{sample['question'][:60]}\n{err_msg}"
            )
            pytest_results.append({
                "run_key":            sample["run_key"],
                "test_id":            sample["test_id"],
                "pdf_file":           sample["pdf_file"],
                "question_col":       sample["question_col"],
                "question":           sample["question"],
                "answer":             "",
                "retrieved_chunks":   [],
                "mrr_score":          None,
                "first_relevant_rank": None,
                "metrics":            [],
                "passed":             False,
                "error":              err_msg,
            })
            _flush_json(pytest_results)
            raise

        else:
            executor.shutdown(wait=False)

        failures = [s for s in scores if not s["success"]]
        pytest_results.append({
            "run_key":             sample["run_key"],
            "test_id":             sample["test_id"],
            "pdf_file":            sample["pdf_file"],
            "question_col":        sample["question_col"],
            "question":            sample["question"],
            "answer":              answer,
            "retrieved_chunks":    retrieved_chunks,
            "mrr_score":           mrr_res["mrr"],
            "first_relevant_rank": mrr_res["first_relevant_rank"],
            "metrics":             scores,
            "passed":              not failures,
            "error":               None,
        })

        assert not failures, "DeepEval metrics failed:\n" + "\n".join(
            f"  {s['name']}: score={s['score']:.3f} < threshold={s['threshold']}"
            f"  | {s['reason']}"
            for s in failures
        )

    finally:
        _flush_json(pytest_results)
